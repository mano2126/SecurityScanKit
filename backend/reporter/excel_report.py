"""
Excel 리포트 생성기 - openpyxl 사용
"""
import os
import asyncio
from datetime import datetime


async def generate(job_id: str, target, results: dict) -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _generate_excel, job_id, target, results)


def _generate_excel(job_id: str, target, results: dict) -> str:
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter

        output_dir = "/tmp/security_reports"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"security_report_{job_id}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        COLORS = {
            "critical": "C62828", "high": "E65100",
            "medium": "F9A825", "low": "2E7D32", "info": "546E7A",
            "header": "1565C0", "light": "F5F7FA"
        }
        severity_labels = {
            "critical": "긴급", "high": "고위험",
            "medium": "중위험", "low": "저위험", "info": "정보"
        }

        def header_style(ws, row, cols, text_list, bg_color="1565C0"):
            for i, (col, text) in enumerate(zip(cols, text_list)):
                cell = ws.cell(row=row, column=col, value=text)
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=Side(style="thin", color="FFFFFF"),
                    right=Side(style="thin", color="FFFFFF"),
                    top=Side(style="thin", color="FFFFFF"),
                    bottom=Side(style="thin", color="FFFFFF"),
                )

        def data_cell(ws, row, col, value, color=None, bold=False, wrap=True, align="left"):
            cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
            cell.font = Font(bold=bold, size=9, color=color or "000000")
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            cell.border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            return cell

        # ── Sheet 1: 요약 ───────────────────────────────────
        ws1 = wb.create_sheet("요약")
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 35
        ws1.column_dimensions["C"].width = 15

        ws1.merge_cells("A1:C1")
        cell = ws1["A1"]
        cell.value = "보안 취약점 점검 결과 보고서"
        cell.font = Font(bold=True, size=14, color="0D1B2A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="E8F0FE")
        ws1.row_dimensions[1].height = 30

        summary = results.get("summary", {})
        scan_info = [
            ("점검 대상 IP", target.ip),
            ("시스템 명칭", target.target_name or "-"),
            ("담당 부서", target.department or "-"),
            ("담당자", target.manager or "-"),
            ("점검 일자", results.get("scan_start", "")[:10]),
            ("전체 위험도", summary.get("risk_level", "N/A")),
            ("위험 점수", f"{summary.get('risk_score', 0)} / 100"),
            ("발견 취약점 수", summary.get("total_vulnerabilities", 0)),
        ]
        for i, (k, v) in enumerate(scan_info, start=3):
            data_cell(ws1, i, 1, k, bold=True)
            data_cell(ws1, i, 2, v)

        # 취약점 통계
        ws1.cell(row=12, column=1, value="심각도별 통계").font = Font(bold=True, size=10)
        stats_headers = ["긴급", "고위험", "중위험", "저위험", "정보", "합계"]
        for i, h in enumerate(stats_headers, 1):
            c = ws1.cell(row=13, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=list(COLORS.values())[i-1] if i <= 5 else "1565C0")
            c.alignment = Alignment(horizontal="center")

        stats_vals = [
            summary.get("critical", 0), summary.get("high", 0),
            summary.get("medium", 0), summary.get("low", 0),
            summary.get("info", 0), summary.get("total_vulnerabilities", 0)
        ]
        for i, v in enumerate(stats_vals, 1):
            c = ws1.cell(row=14, column=i, value=v)
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True, size=11)

        wb.save(output_path)
        return output_path

    except ImportError:
        output_dir = "/tmp/security_reports"
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"security_report_{job_id}.txt")
        return output_path
