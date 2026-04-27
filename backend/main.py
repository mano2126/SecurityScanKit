import logging
import logging.config
import urllib3

# ── WS 프레임 로그 차단 필터 ─────────────────────────────────────
# websockets 라이브러리가 루트 로거를 통해 직접 찍는 DEBUG 로그 차단
class _WSFrameFilter(logging.Filter):
    """> TEXT, < TEXT, PING, PONG, connection is 등 WS 프레임 로그 완전 차단"""
    _STARTS = (
        "> TEXT", "< TEXT", "> BINARY", "< BINARY",
        "> PING", "< PONG", "> CLOSE", "< CLOSE",
        "= connection", "x half-", "% sending", "% received",
        "< GET /ws", "< host:", "< upgrade", "< sec-web",
        "< connection:", "< pragma", "< cache-control",
        "< user-agent", "< accept", "< origin",
        "> HTTP/1.1 101", "> Upgrade:", "> Connection:",
        "> Sec-WebSocket", "> date:", "> server:",
    )
    def filter(self, record):
        try:
            msg = record.getMessage()
            return not any(msg.startswith(p) for p in self._STARTS)
        except Exception:
            return True

_ws_frame_filter = _WSFrameFilter()

def _apply_ws_filter():
    """모든 로거의 모든 핸들러에 WS 프레임 필터 적용"""
    import logging as _lg
    manager = _lg.Logger.manager
    # 루트 포함 등록된 모든 로거
    all_loggers = [_lg.root] + [
        v for v in manager.loggerDict.values()
        if isinstance(v, _lg.Logger)
    ]
    for _l in all_loggers:
        for _h in _l.handlers:
            if not any(isinstance(f, _WSFrameFilter) for f in _h.filters):
                _h.addFilter(_ws_frame_filter)

# 모듈 로드 시점에 즉시 적용
_apply_ws_filter()

_WEBSOCKET_SILENCED = True
for _s in [
    "websockets", "websockets.server", "websockets.client",
    "websockets.protocol", "websockets.legacy",
    "websockets.legacy.server", "websockets.legacy.client",
    "websockets.legacy.protocol",
    "uvicorn.protocols", "uvicorn.protocols.websockets",
    "uvicorn.protocols.websockets.websockets_impl",
    "uvicorn.protocols.websockets.wsproto_impl",
    "uvicorn.protocols.http", "uvicorn.protocols.http.h11_impl",
    "uvicorn.protocols.http.httptools_impl",
    "h11", "asyncio",
]:
    _sl = logging.getLogger(_s)
    _sl.setLevel(logging.CRITICAL)
    _sl.handlers = [logging.NullHandler()]
    _sl.propagate = False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# 로그 설정 — websockets 노이즈 차단
try:
    from log_config import LOG_CONFIG
    logging.config.dictConfig(LOG_CONFIG)
except Exception:
    pass
"""
SecurityScanKit v1.0 — FastAPI 백엔드 메인
모든 API 엔드포인트가 실제 SQLite DB와 연동됨
"""
import uuid, asyncio, hashlib, json, os
from datetime import datetime
from typing import List, Optional
from pathlib import Path

from fastapi import FastAPI, BackgroundTasks, HTTPException, Depends, UploadFile, File, Form, WebSocket, WebSocketDisconnect, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from config_loader import get_config
from db import (
    init_db, get_db,
    Asset, ScanJob, Finding, Alert, CVERecord, NewsItem,
    Notification,
    create_asset, get_asset, get_asset_by_ip, get_all_assets,
    update_asset, update_asset_risk_score, delete_asset, bulk_upsert_assets,
    create_scan_job, get_scan_job, update_scan_job,
    get_recent_scan_jobs, get_asset_scan_history,
    save_finding, get_all_findings, get_findings_by_asset,
    get_repeat_findings, resolve_finding, get_finding_stats,
    create_alert, get_alerts, mark_alerts_read, get_unread_alert_count,
    upsert_cve, get_recent_cves, get_cves_affecting_assets, save_cve_impact,
    save_news, get_news,
    save_upload_history, get_upload_history,
    get_alert_configs, update_alert_config,
    get_dashboard_stats,
)
from scanner.port_scanner    import PortScanner
from scanner.web_scanner     import WebScanner
from scanner.ssl_scanner     import SSLScanner
from scanner.db_scanner      import DBScanner
from scanner.network_scanner import NetworkScanner
from ai_analyzer             import AIAnalyzer
from reporter                import PDFReporter, ExcelReporter

cfg = get_config()
cfg.print_summary()

app = FastAPI(title="SecurityScanKit API", version="1.0.0", docs_url="/docs")

app.add_middleware(CORSMiddleware, allow_origins=["*"],
    allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ── 접속자 추적 시스템 ────────────────────────────────────────
import collections as _col

_access_log: _col.deque = _col.deque(maxlen=500)
_active_sessions: dict = {}

# Windows 소켓 누수 방지 — BaseHTTPMiddleware 대신 순수 ASGI 미들웨어 사용
class AccessTracker:
    def __init__(self, app):
        self._app = app

    async def __call__(self, scope, receive, send):
        if scope["type"] not in ("http", "websocket"):
            await self._app(scope, receive, send)
            return

        path = scope.get("path", "")
        if path in {"/docs", "/openapi.json", "/favicon.ico"} or path.startswith("/static"):
            await self._app(scope, receive, send)
            return

        # IP 추출
        client = scope.get("client")
        ip = client[0] if client else "unknown"
        for hdr_name, hdr_val in scope.get("headers", []):
            if hdr_name == b"x-forwarded-for":
                ip = hdr_val.decode().split(",")[0].strip(); break
            if hdr_name == b"x-real-ip":
                ip = hdr_val.decode().strip(); break

        # User-Agent
        ua = ""
        for hdr_name, hdr_val in scope.get("headers", []):
            if hdr_name == b"user-agent":
                ua = hdr_val.decode()[:200]; break

        def parse_browser(u):
            ul = u.lower()
            if "edg/" in ul:    return "Edge"
            if "chrome/" in ul: return "Chrome"
            if "firefox/" in ul:return "Firefox"
            if "safari/" in ul: return "Safari"
            return "Other" if u else "Unknown"

        def parse_os(u):
            ul = u.lower()
            if "windows nt 10" in ul: return "Windows 10/11"
            if "windows" in ul:        return "Windows"
            if "mac os" in ul:         return "macOS"
            if "linux" in ul:          return "Linux"
            if "android" in ul:        return "Android"
            if "iphone" in ul:         return "iOS"
            return "Unknown"

        now = datetime.now()
        status_code = [200]

        async def send_wrapper(message):
            if message["type"] == "http.response.start":
                status_code[0] = message.get("status", 200)
            await send(message)

        start = datetime.now()
        try:
            await self._app(scope, receive, send_wrapper)
        finally:
            if path.startswith("/api") or path.startswith("/ws"):
                lat = round((datetime.now() - start).total_seconds() * 1000)
                method = scope.get("method", "WS")
                entry = {
                    "ts": now.strftime("%Y-%m-%d %H:%M:%S"),
                    "ip": ip, "method": method, "path": path,
                    "status": status_code[0], "latency_ms": lat,
                    "browser": parse_browser(ua), "os": parse_os(ua), "ua": ua,
                }
                _access_log.appendleft(entry)
                sess = _active_sessions.get(ip, {})
                _active_sessions[ip] = {
                    "ip": ip, "browser": parse_browser(ua), "os": parse_os(ua),
                    "last_seen": now.strftime("%Y-%m-%d %H:%M:%S"),
                    "last_path": path, "ua": ua,
                    "first_seen": sess.get("first_seen", now.strftime("%Y-%m-%d %H:%M:%S")),
                    "req_count": sess.get("req_count", 0) + 1,
                }

app.add_middleware(AccessTracker)

@app.get("/api/admin/visitors")
def get_visitors():
    """접속자 현황 — 활성 세션 + 최근 접속 이력"""
    now = datetime.now()
    # 10분 이내 접속을 '현재 접속중'으로 간주
    active = []
    for ip, s in _active_sessions.items():
        try:
            last = datetime.strptime(s["last_seen"], "%Y-%m-%d %H:%M:%S")
            diff = (now - last).total_seconds()
            active.append({**s, "minutes_ago": round(diff/60, 1),
                           "is_active": diff < 600})
        except Exception:
            pass
    active.sort(key=lambda x: x["last_seen"], reverse=True)
    return {
        "active_sessions": active,
        "total_unique_ips": len(_active_sessions),
        "recent_logs": list(_access_log)[:100],
        "total_requests": len(_access_log),
    }


import collections
_scan_logs: dict = {}           # job_id → deque(maxlen=500)
_ws_clients: dict = {}          # job_id → set of WebSocket

def push_log(job_id: str, level: str, tag: str, msg: str):
    """스캔 로그 저장 + 연결된 WebSocket 클라이언트에 전송"""
    entry = {
        "time":    datetime.now().strftime("%H:%M:%S"),
        "level":   level,
        "tag":     tag,
        "message": msg,
    }
    if job_id not in _scan_logs:
        _scan_logs[job_id] = collections.deque(maxlen=500)
    _scan_logs[job_id].append(entry)
    # 연결된 클라이언트에 즉시 전송
    if job_id in _ws_clients:
        import asyncio
        dead = set()
        for ws in _ws_clients[job_id]:
            try:
                asyncio.create_task(ws.send_json(entry))
            except Exception:
                dead.add(ws)
        _ws_clients[job_id] -= dead

@app.on_event("startup")
async def startup():
    init_db()

    # ── WS 프레임 노이즈 완전 제거 ─────────────────────────────
    # uvicorn이 내부적으로 websockets 라이브러리 DEBUG 로그를 찍는데
    # 이것이 우리 로그 핸들러와 무한 루프를 만듦 → WARNING 강제 설정
    import logging as _logging
    _ws_suppress = [
        "websockets", "websockets.client", "websockets.server",
        "websockets.protocol", "websockets.legacy",
        "websockets.legacy.server", "websockets.legacy.client",
        "uvicorn.protocols", "uvicorn.protocols.websockets",
        "uvicorn.protocols.http",
    ]
    for _ln in _ws_suppress:
        _l2 = _logging.getLogger(_ln)
        _l2.setLevel(_logging.WARNING)
        _l2.propagate = False
        # 기존 핸들러 제거 후 재추가 방지
        _l2.handlers = [h for h in _l2.handlers if not isinstance(h, _AppLogHandler)]

    # startup 시점에 uvicorn이 등록한 모든 핸들러에 필터 재적용
    _apply_ws_filter()
    print("[APP] 서버 시작 완료")
    app_logger.info("[APP] SecurityScanKit 백엔드 시작")
    # 서버 시작 시 즉시 1회 + 이후 24시간마다 자동 수집
    pass  # 보안 뉴스 자동 수집 비활성화
    # asyncio.create_task(_schedule_intel_collect())

async def _schedule_intel_collect():
    """위협 인텔리전스 자동 수집 — 비활성화됨"""
    pass  # 보안 뉴스 수집 기능 비활성화

# ── Pydantic 모델 ──────────────────────────────────────────────
class AssetCreate(BaseModel):
    name: str; ip: str
    asset_type:  Optional[str] = "웹서버"
    environment: Optional[str] = "Production"
    department:  Optional[str] = ""
    manager:     Optional[str] = ""
    priority:    Optional[str] = "medium"
    scan_types:  Optional[str] = "port,web,ssl"
    http_port:   Optional[int] = 80
    https_port:  Optional[int] = 443
    db_type:     Optional[str] = None
    db_port:     Optional[int] = None
    note:        Optional[str] = ""

class AssetUpdate(BaseModel):
    name:        Optional[str] = None
    asset_type:  Optional[str] = None
    environment: Optional[str] = None
    department:  Optional[str] = None
    manager:     Optional[str] = None
    priority:    Optional[str] = None
    scan_types:  Optional[str] = None
    note:        Optional[str] = None

class ScanRequest(BaseModel):
    asset_ids:  List[str]
    scan_types: Optional[str] = "port,web,ssl"

class FindingResolve(BaseModel):
    resolved_by: str
    note: Optional[str] = ""

class AlertConfigUpdate(BaseModel):
    is_active:       Optional[bool] = None
    threshold_value: Optional[int]  = None
    channels:        Optional[str]  = None

# ── 직렬화 헬퍼 ───────────────────────────────────────────────
def _f(f: Finding) -> dict:
    d = {c.name: getattr(f, c.name) for c in f.__table__.columns}
    for k in ("first_seen","last_seen","created_at","resolved_at"):
        if d.get(k): d[k] = d[k].isoformat()
    if f.asset:
        d["asset_name"] = f.asset.name
        d["asset_ip"]   = f.asset.ip
    return d

def _a(a: Asset) -> dict:
    d = {c.name: getattr(a, c.name) for c in a.__table__.columns}
    for k in ("last_scan","created_at","updated_at"):
        if d.get(k): d[k] = d[k].isoformat()
    return d

def _j(j: ScanJob) -> dict:
    d = {c.name: getattr(j, c.name) for c in j.__table__.columns}
    for k in ("started_at","completed_at","created_at"):
        if d.get(k): d[k] = d[k].isoformat()
    if j.asset:
        d["asset_name"] = j.asset.name
        d["asset_ip"]   = j.asset.ip
    return d

# ══════════════════════════════════════════════════════════════
# ENDPOINTS
# ══════════════════════════════════════════════════════════════

@app.get("/api/health")
def health():
    return {"status":"ok","version":"1.0.0","time":datetime.now().isoformat(),
            "os":cfg.os,"nmap":cfg.nmap_available}

# ── 대시보드 ──────────────────────────────────────────────────
@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db)):
    stats    = get_dashboard_stats(db)
    alerts   = get_alerts(db, limit=5)
    assets   = sorted(get_all_assets(db), key=lambda x: x.risk_score, reverse=True)[:5]
    repeats  = get_repeat_findings(db, threshold=cfg.repeat_threshold)[:5]
    news     = get_news(db, limit=5)
    return {
        "stats": stats,
        "top_assets": [_a(a) for a in assets],
        "recent_alerts": [
            {"id":a.id,"title":a.title,"severity":a.severity,
             "alert_type":a.alert_type,"is_read":a.is_read,
             "created_at":a.created_at.isoformat()} for a in alerts
        ],
        "repeat_findings": [_f(f) for f in repeats],
        "recent_news": [
            {"id":n.id,"source":n.source,"source_tag":n.source_tag,
             "title":n.title,"severity":n.severity,"affects_assets":n.affects_assets,
             "published_at":n.published_at.isoformat() if n.published_at else None}
            for n in news
        ],
    }

# ── 자산 ──────────────────────────────────────────────────────
@app.get("/api/assets")
def list_assets(db: Session = Depends(get_db)):
    return [_a(a) for a in get_all_assets(db)]

@app.post("/api/assets", status_code=201)
def add_asset(data: AssetCreate, db: Session = Depends(get_db)):
    if get_asset_by_ip(db, data.ip):
        raise HTTPException(400, f"IP {data.ip} 는 이미 등록되어 있습니다")
    return _a(create_asset(db, data.model_dump()))

@app.get("/api/assets/{asset_id}")
def get_asset_detail(asset_id: str, db: Session = Depends(get_db)):
    a = get_asset(db, asset_id)
    if not a: raise HTTPException(404, "자산을 찾을 수 없습니다")
    d = _a(a)
    d["findings"] = [_f(f) for f in get_findings_by_asset(db, asset_id, "open")]
    d["scan_history"] = [_j(j) for j in get_asset_scan_history(db, asset_id)]
    return d

@app.put("/api/assets/{asset_id}")
def edit_asset(asset_id: str, data: AssetUpdate, db: Session = Depends(get_db)):
    updates = {k:v for k,v in data.model_dump().items() if v is not None}
    a = update_asset(db, asset_id, updates)
    if not a: raise HTTPException(404, "자산을 찾을 수 없습니다")
    return _a(a)

@app.delete("/api/assets/{asset_id}")
def remove_asset(asset_id: str, db: Session = Depends(get_db)):
    if not delete_asset(db, asset_id): raise HTTPException(404, "자산을 찾을 수 없습니다")
    return {"message": "삭제 완료"}

@app.post("/api/assets/upload")
async def upload_assets(
    file: UploadFile = File(...),
    uploaded_by: str = Form(default="admin"),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith((".xlsx",".xls")):
        raise HTTPException(400, ".xlsx 파일만 업로드 가능합니다")
    contents = await file.read()
    file_hash = hashlib.sha256(contents).hexdigest()
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[5]]
        col_map = {
            "IP 주소 *":"ip","시스템 명칭 *":"name","시스템 유형 *":"asset_type",
            "운영환경 *":"environment","점검유형 *":"scan_types",
            "HTTP포트":"http_port","HTTPS포트":"https_port","DB유형":"db_type",
            "DB포트":"db_port","담당부서":"department","담당자명":"manager",
            "우선순위":"priority","비고/특이사항":"note"
        }
        priority_map = {
            "긴급(Critical)":"critical","높음(High)":"high",
            "보통(Medium)":"medium","낮음(Low)":"low"
        }
        assets_data = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row[1]: continue
            data = {}
            for ci, hdr in enumerate(headers):
                key = col_map.get(hdr)
                if key and ci < len(row) and row[ci] is not None:
                    val = str(row[ci]).strip()
                    if key in ("http_port","https_port","db_port"):
                        try: data[key] = int(val)
                        except: pass
                    else:
                        data[key] = val
            if data.get("ip") and data.get("name"):
                data["priority"] = priority_map.get(data.get("priority",""), "medium")
                assets_data.append(data)
        upload_rec = save_upload_history(db, file.filename, uploaded_by, len(assets_data), file_hash)
        count = bulk_upsert_assets(db, assets_data, upload_rec.id)
        return {"message":f"{count}개 자산 등록/갱신","asset_count":count,
                "upload_id":upload_rec.id,"filename":file.filename}
    except Exception as e:
        raise HTTPException(500, f"파일 처리 오류: {str(e)}")

@app.get("/api/assets/upload/history")
def upload_history_list(db: Session = Depends(get_db)):
    return [
        {"id":u.id,"filename":u.filename,"uploaded_by":u.uploaded_by,
         "asset_count":u.asset_count,"version":u.version,"status":u.status,
         "created_at":u.created_at.isoformat()} for u in get_upload_history(db)
    ]

# ── 점검 엔진 ─────────────────────────────────────────────────
async def _run_scan(job_id: str, asset_id: str, scan_types_str: str):
    from db.schema import SessionLocal as SL
    db = SL()
    try:
        asset = get_asset(db, asset_id)
        if not asset: return
        update_scan_job(db, job_id, {"status":"running","started_at":datetime.now(),"current_step":"초기화 중..."})

        scan_types = [s.strip() for s in scan_types_str.split(",")]
        all_findings: dict = {}
        total = len(scan_types); done = 0

        def upd(msg, level="INFO"):
            nonlocal done
            update_scan_job(db, job_id, {"progress":int((done/total)*85),"current_step":msg})
            push_log(job_id, level, "SCAN", msg)

        if "port" in scan_types:
            upd("포트 스캔 시작 — 위험 포트 목록 로딩")
            push_log(job_id, "INFO", "SCAN", "[포트] STEP1 위험 포트 연결 시도 중...")
            r_port_partial = await asyncio.get_event_loop().run_in_executor(
                None, lambda: None)  # 잠깐 await → 로그 flush
            push_log(job_id, "INFO", "SCAN", "[포트] STEP2 서비스 식별 중 (SSH/RDP/HTTP 등 버전 확인)")
            await asyncio.sleep(0.1)  # STEP2 로그 먼저 전달
            r = await PortScanner(asset.ip).run()
            push_log(job_id, "INFO", "SCAN", "[포트] STEP3 위험 포트 판정 중 (금융보안원 기준 적용)")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[포트] STEP4 결과 저장 완료")
            all_findings["port"] = r.get("vulnerabilities",[]); done+=1
        if "web" in scan_types:
            upd("웹 취약점 점검 시작 — HTTP 응답 분석")
            push_log(job_id, "INFO", "SCAN", "[웹] STEP1 HTTP 응답 헤더 검사 중 (X-Frame-Options, CSP 등)")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[웹] STEP2 민감 경로 노출 검사 중 (/.env, /admin 등 18종)")
            await asyncio.sleep(0.1)
            r = await WebScanner(asset.ip, asset.http_port or 80, asset.https_port or 443).run()
            push_log(job_id, "INFO", "SCAN", "[웹] STEP3 HTTP→HTTPS 리다이렉트 전환 검사 중")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[웹] STEP4 결과 저장 완료")
            all_findings["web"] = r.get("vulnerabilities",[]); done+=1
        if "ssl" in scan_types:
            upd("SSL/TLS 점검 시작 — 인증서 연결 시도")
            push_log(job_id, "INFO", "SCAN", "[SSL] STEP1 SSL 연결 시도 중 (443 포트) — 응답 대기 중...")
            push_log(job_id, "INFO", "SCAN", "[SSL] STEP2 인증서 유효성 확인 중 (만료일/발급기관/자체서명)")

            # SSL 점검은 시간이 걸림 — 진행 중 알림 태스크 실행
            async def _ssl_progress():
                msgs = [
                    "[SSL] 🔒 TLS 핸드셰이크 진행 중... (정상 소요 시간: 10~30초)",
                    "[SSL] 🔒 프로토콜 버전 협상 중... TLSv1.3 / TLSv1.2 확인",
                    "[SSL] 🔒 인증서 체인 검증 중... 잠시 기다려 주세요",
                    "[SSL] 🔒 암호화 스위트 목록 수집 중...",
                ]
                for i, msg in enumerate(msgs):
                    await asyncio.sleep(6)
                    push_log(job_id, "INFO", "SCAN", msg)

            progress_task = asyncio.create_task(_ssl_progress())
            try:
                r = await SSLScanner(asset.ip, asset.https_port or 443).run()
            finally:
                progress_task.cancel()

            push_log(job_id, "INFO", "SCAN", "[SSL] STEP3 프로토콜 버전 검사 완료 (TLSv1.0/1.1 취약 여부)")
            push_log(job_id, "INFO", "SCAN", "[SSL] STEP4 암호화 알고리즘 검사 완료 (RC4/DES/3DES 등)")
            all_findings["ssl"] = r.get("vulnerabilities",[]); done+=1
        if "db" in scan_types and asset.db_type:
            upd("DB 점검 시작 — 포트 접근 가능 여부 확인")
            push_log(job_id, "INFO", "SCAN", f"[DB] STEP1 DB 포트 스캔 중 ({asset.db_type.upper()} {asset.db_port or '기본포트'})")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[DB] STEP2 외부망 직접 접근 가능성 판정 중")
            await asyncio.sleep(0.1)
            r = await DBScanner(asset.ip, asset.db_type, asset.db_port).run()
            push_log(job_id, "INFO", "SCAN", "[DB] STEP3 기본 계정 활성화 여부 확인 중")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[DB] STEP4 결과 저장 완료")
            all_findings["db"] = r.get("vulnerabilities",[]); done+=1
        if "network" in scan_types:
            upd("네트워크 장비 점검 시작 — Telnet 포트 확인")
            push_log(job_id, "INFO", "SCAN", "[네트워크] STEP1 Telnet 포트(23) 개방 여부 검사 중")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[네트워크] STEP2 SNMP Community String 기본값 사용 검사 중")
            await asyncio.sleep(0.1)
            r = await NetworkScanner(asset.ip).run()
            push_log(job_id, "INFO", "SCAN", "[네트워크] STEP3 기본 자격증명(admin/admin) 시도 중")
            await asyncio.sleep(0.1)
            push_log(job_id, "INFO", "SCAN", "[네트워크] STEP4 결과 저장 완료")

        update_scan_job(db, job_id, {"progress":87,"current_step":"취약점 저장 중..."})
        crit=high=med=low=info_c=0
        for scan_type, vulns in all_findings.items():
            for v in vulns:
                sev = v.get("severity","info")
                saved = save_finding(db, {
                    "asset_id":asset_id,"scan_job_id":job_id,
                    "vuln_id":v.get("id",str(uuid.uuid4())[:8]),
                    "title":v.get("title",""),"description":v.get("description",""),
                    "recommendation":v.get("recommendation",""),
                    "severity":sev,"cvss_score":v.get("cvss_score",0.0),
                    "scan_type":scan_type,"port":v.get("port"),
                    "service":v.get("service",""),"regulation":v.get("reference",""),
                    "raw_output":str(v)[:2000],
                })
                push_log(job_id, "WARN" if sev in ("critical","high") else "INFO",
                    sev.upper(), f"발견: [{sev.upper()}] {v.get('title','')[:60]}")
                if saved.repeat_count >= cfg.repeat_threshold:
                    push_log(job_id, "WARN", "REPEAT", f"반복 취약점 {saved.repeat_count}회: {saved.title[:50]}")
                    create_alert(db,"repeat_vuln",sev,
                        f"반복 취약점: {saved.title[:80]}",
                        f"{asset.name}({asset.ip}) {saved.repeat_count}회 연속 발견",
                        asset_id=asset_id, finding_id=saved.id)
                if sev=="critical": crit+=1
                elif sev=="high": high+=1
                elif sev=="medium": med+=1
                elif sev=="low": low+=1
                else: info_c+=1

        if crit > 0:
            create_alert(db,"critical_found","critical",
                f"긴급 취약점 {crit}건: {asset.name}",
                f"{asset.ip} 점검 결과 Critical {crit}건", asset_id=asset_id)

        update_scan_job(db, job_id, {"progress":93,"current_step":"AI 분석 중..."})
        if cfg.ai_enabled and (crit+high) > 0:
            try:
                ai_r = await AIAnalyzer().analyze(all_findings)
                from db.schema import Finding as F
                fc = db.query(F).filter(F.scan_job_id==job_id, F.severity=="critical").first()
                if fc and ai_r:
                    fc.ai_analysis = str(ai_r.get("executive_summary",""))[:2000]
                    db.commit()
            except: pass

        # ── 이번 점검에서 발견 안 된 기존 취약점 → 자동 resolved ──
        # 이번 점검에서 발견된 vuln_id 목록
        found_vuln_ids = set()
        for scan_type, vulns in all_findings.items():
            for v in vulns:
                found_vuln_ids.add(v.get("id",""))

        # 이번 점검 대상 scan_type 목록
        scanned_types = [t for t in ["port","web","ssl","db","network"] if t in scan_types]

        # 해당 자산의 open 취약점 중 이번 점검 유형에 속하는데 발견 안 된 것 → resolved
        from db.schema import Finding as _Finding
        auto_resolved = 0
        old_open = db.query(_Finding).filter(
            _Finding.asset_id == asset_id,
            _Finding.status   != "resolved",
            _Finding.scan_type.in_(scanned_types),
        ).all()
        for f in old_open:
            if f.vuln_id not in found_vuln_ids:
                f.status      = "resolved"
                f.resolved_at = datetime.now()
                f.resolved_by = "AUTO:재점검_미발견"
                auto_resolved += 1
        if auto_resolved > 0:
            db.commit()
            push_log(job_id, "INFO", "RESOLVE",
                f"재점검 미발견 취약점 {auto_resolved}건 자동 해소 처리")

        update_scan_job(db, job_id, {"progress":97,"current_step":"위험 점수 계산 중..."})
        update_asset_risk_score(db, asset_id)
        update_asset(db, asset_id, {"last_scan": datetime.now()})

        now = datetime.now()
        job = get_scan_job(db, job_id)
        dur = (now - job.started_at).total_seconds() if job and job.started_at else 0
        push_log(job_id, "INFO", "DONE",
            f"점검 완료 — Critical:{crit} High:{high} Medium:{med} Low:{low} ({round(dur,1)}초)")
        update_scan_job(db, job_id, {
            "status":"completed","progress":100,"current_step":"완료",
            "completed_at":now,"duration_sec":dur,
            "crit_count":crit,"high_count":high,"med_count":med,"low_count":low,"info_count":info_c,
        })
    except Exception as e:
        import traceback
        err = type(e).__name__ + ": " + str(e) + "\n" + traceback.format_exc()
        update_scan_job(db, job_id, {"status":"failed","error_msg":err[:2000],"progress":0})
        print("[SCAN ERROR] " + err)
    finally:
        db.close()

@app.post("/api/scan/start")
async def start_scan(req: ScanRequest, bg: BackgroundTasks, db: Session = Depends(get_db)):
    jobs = []
    for aid in req.asset_ids:
        a = get_asset(db, aid)
        if not a: continue
        job = create_scan_job(db, aid, req.scan_types)
        bg.add_task(_run_scan, job.id, aid, req.scan_types)
        jobs.append({"job_id":job.id,"asset_id":aid,"asset_name":a.name})
    return {"message":f"{len(jobs)}개 자산 점검 시작","jobs":jobs}

@app.get("/api/scan/status/{job_id}")
def scan_status(job_id: str, db: Session = Depends(get_db)):
    j = get_scan_job(db, job_id)
    if not j: raise HTTPException(404,"점검 작업을 찾을 수 없습니다")
    return _j(j)

@app.get("/api/scan/history")
def scan_history(db: Session = Depends(get_db)):
    return [_j(j) for j in get_recent_scan_jobs(db, 50)]

# ── 취약점 ──────────────────────────────────────────────────
@app.get("/api/findings")
def list_findings(severity:Optional[str]=None, status:Optional[str]=None,
                  scan_type:Optional[str]=None, repeat_only:bool=False,
                  db: Session = Depends(get_db)):
    return [_f(f) for f in get_all_findings(db, severity=severity, status=status,
                                             scan_type=scan_type, repeat_only=repeat_only)]

@app.get("/api/findings/stats")
def finding_stats(db: Session = Depends(get_db)):
    return get_finding_stats(db)

@app.get("/api/findings/repeat")
def repeat_findings(db: Session = Depends(get_db)):
    return [_f(f) for f in get_repeat_findings(db, threshold=cfg.repeat_threshold)]

@app.post("/api/findings/{finding_id}/resolve")
def resolve(finding_id: str, data: FindingResolve, db: Session = Depends(get_db)):
    f = resolve_finding(db, finding_id, data.resolved_by, data.note)
    if not f: raise HTTPException(404,"취약점을 찾을 수 없습니다")
    return {"message":"조치 완료"}

@app.delete("/api/findings/{finding_id}")
def delete_finding(finding_id: str, db: Session = Depends(get_db)):
    """취약점 단건 삭제"""
    from db.schema import Finding
    f = db.query(Finding).filter(Finding.id == finding_id).first()
    if not f: raise HTTPException(404, "취약점을 찾을 수 없습니다")
    db.delete(f)
    db.commit()
    return {"message": "삭제 완료"}

@app.delete("/api/findings")
def delete_findings_bulk(ids: list[str], db: Session = Depends(get_db)):
    """취약점 일괄 삭제"""
    from db.schema import Finding
    db.query(Finding).filter(Finding.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"{len(ids)}건 삭제 완료"}

# ── 알람 ────────────────────────────────────────────────────
@app.get("/api/alerts")
def list_alerts(unread_only: bool = False, db: Session = Depends(get_db)):
    return [{"id":a.id,"alert_type":a.alert_type,"severity":a.severity,"title":a.title,
             "message":a.message,"asset_id":a.asset_id,"is_read":a.is_read,
             "created_at":a.created_at.isoformat()} for a in get_alerts(db, unread_only)]

@app.get("/api/alerts/count")
def alert_count(db: Session = Depends(get_db)):
    return {"unread": get_unread_alert_count(db)}

@app.post("/api/alerts/read")
def read_alerts(alert_ids: List[str], db: Session = Depends(get_db)):
    return {"read": mark_alerts_read(db, alert_ids)}

@app.get("/api/alerts/config")
def alert_configs(db: Session = Depends(get_db)):
    return [{"id":c.id,"alert_type":c.alert_type,"label":c.label,"condition":c.condition,
             "channels":c.channels,"is_active":c.is_active,"threshold_value":c.threshold_value}
            for c in get_alert_configs(db)]

@app.put("/api/alerts/config/{alert_type}")
def update_alert_cfg(alert_type: str, data: AlertConfigUpdate, db: Session = Depends(get_db)):
    updates = {k:v for k,v in data.model_dump().items() if v is not None}
    c = update_alert_config(db, alert_type, updates)
    if not c: raise HTTPException(404,"알람 설정을 찾을 수 없습니다")
    return {"message":"업데이트 완료"}

# ── CVE ─────────────────────────────────────────────────────
@app.get("/api/cve")
def list_cve(days: int = 30, limit: int = 200, severity: str = "", db: Session = Depends(get_db)):
    from db.schema import CVERecord, CVEImpact
    from sqlalchemy import inspect as sa_inspect
    # CVERecord 테이블 존재 여부 확인
    inspector = sa_inspect(db.bind)
    tables = inspector.get_table_names()
    if "cve_records" in tables:
        q = db.query(CVERecord)
        if days:
            from datetime import timedelta
            cutoff = datetime.now() - timedelta(days=days)
            q = q.filter(CVERecord.created_at >= cutoff)
        if severity:
            q = q.filter(CVERecord.severity == severity)
        cves = q.order_by(CVERecord.cvss_score.desc()).limit(limit).all()
        return [_cve_r(c) for c in cves]
    # fallback: 구버전 CVEImpact 스키마
    result = []
    for c in get_recent_cves(db, days=days):
        result.append({"id":c.id,"cve_id":c.cve_id or "","title":c.title or "",
                       "description":c.description or "","cvss_score":c.cvss_score,
                       "severity":c.severity or "","published_at":str(c.published_date or ""),
                       "affected_products":c.affected_products or "","source":"NVD",
                       "url":f"https://nvd.nist.gov/vuln/detail/{c.cve_id}" if c.cve_id else "",
                       "cwe":"","kev_due_date":"","kev_action":"","cvss_vector":"",
                       "created_at":str(c.created_at or "")})
    return result

@app.get("/api/cve/stats")
def cve_stats(db: Session = Depends(get_db)):
    from db.schema import CVERecord
    total  = db.query(CVERecord).count()
    crit   = db.query(CVERecord).filter(CVERecord.severity=="critical").count()
    high   = db.query(CVERecord).filter(CVERecord.severity=="high").count()
    kev    = db.query(CVERecord).filter(CVERecord.is_kev==True).count()
    week   = datetime.now() - __import__("datetime").timedelta(days=7)
    new7   = db.query(CVERecord).filter(CVERecord.created_at>=week).count()
    return {"total":total,"critical":crit,"high":high,"kev":kev,"new_7d":new7}

def _cve_r(c) -> dict:
    cve_id = c.id  # CVERecord의 PK가 CVE-XXXX-XXXXX 형식
    return {
        "id":c.id,"cve_id":cve_id,
        "title":f"{cve_id}: {(c.description or '')[:120]}",
        "description":c.description or "",
        "cvss_score":c.cvss_score or 0.0,
        "cvss_vector":c.cvss_vector or "",
        "severity":c.severity or "medium",
        "published_at":c.published_date.isoformat() if c.published_date else "",
        "affected_products":c.affected_products or "[]",
        "source":"CISA KEV" if c.is_kev else "NVD",
        "url":f"https://nvd.nist.gov/vuln/detail/{cve_id}",
        "cwe":"",
        "kev_due_date":c.kev_added_date.isoformat() if c.kev_added_date else "",
        "kev_action":c.patch_info or "",
        "created_at":c.created_at.isoformat() if c.created_at else "",
    }

# ── 뉴스 ────────────────────────────────────────────────────
@app.get("/api/news")
def list_news(source: Optional[str] = None, severity: str = "", limit: int = 200, db: Session = Depends(get_db)):
    items = get_news(db, source=source, limit=limit)
    if severity:
        items = [n for n in items if n.severity == severity]
    return [_news_r(n) for n in items]

@app.get("/api/news/stats")
def news_stats(db: Session = Depends(get_db)):
    from db.schema import NewsItem
    total = db.query(NewsItem).count()
    crit  = db.query(NewsItem).filter(NewsItem.severity=="critical").count()
    high  = db.query(NewsItem).filter(NewsItem.severity=="high").count()
    week  = datetime.now() - __import__("datetime").timedelta(days=7)
    new7  = db.query(NewsItem).filter(NewsItem.created_at>=week).count()
    sources = {}
    for row in db.query(NewsItem.source, NewsItem.id).all():
        sources[row[0]] = sources.get(row[0], 0) + 1
    return {"total":total,"critical":crit,"high":high,"new_7d":new7,"sources":sources}

def _news_r(n) -> dict:
    return {
        "id":n.id,"source":n.source or "","source_tag":n.source_tag or "",
        "title":n.title or "","summary":n.summary or "","url":n.url or "",
        "severity":n.severity or "medium",
        "published_at":n.published_at.isoformat() if n.published_at else "",
        "affects_assets":n.affects_assets or False,
        "created_at":n.created_at.isoformat() if n.created_at else "",
    }

@app.post("/api/news/fetch")
async def fetch_news(bg: BackgroundTasks):
    from core.intel_collector import collect_all_with_progress
    if _collect_progress.get("running"):
        return {"message":"이미 수집 중입니다", "running": True}
    config = _collect_config.copy()  # 항상 현재 설정 사용 (기본값 포함)
    bg.add_task(collect_all_with_progress, config, _collect_progress)
    return {"message":"위협 인텔리전스 수집 시작", "running": False}

# ── 컴플라이언스 ─────────────────────────────────────────────
@app.get("/api/compliance")
def compliance(db: Session = Depends(get_db)):
    findings = get_all_findings(db, status="open")
    compliance_map = {
        "ISMS-P":["web","ssl"],
        "전자금융감독규정":["port","ssl","network"],
        "금융보안원 가이드":["port","db","network"],
        "PCI-DSS":["web","ssl","db"],
        "ISO 27001":["port","web","ssl"],
        "NIST CSF":["port","web","ssl","db","network"],
    }
    base = {"ISMS-P":102,"전자금융감독규정":45,"금융보안원 가이드":80,"PCI-DSS":64,"ISO 27001":93,"NIST CSF":108}
    result = []
    for std, types in compliance_map.items():
        related = [f for f in findings if f.scan_type in types]
        critical_r = [f for f in related if f.severity in ("critical","high")]
        score = max(0, 100 - min(len(critical_r)*3 + len(related), 40))
        result.append({"standard":std,"score":score,"total_items":base.get(std,100),
                       "issues":len(critical_r),"related_vulns":len(related)})
    return result

# ── 리포트 ───────────────────────────────────────────────────
@app.post("/api/report/generate")
async def gen_report(report_type: str = "executive", db: Session = Depends(get_db)):
    findings_all = get_all_findings(db, limit=500)
    assets   = get_all_assets(db)
    stats    = get_finding_stats(db)
    job_id   = str(uuid.uuid4())[:8].upper()
    data = {
        "generated_at": datetime.now().isoformat(),
        "report_type":  report_type,
        "stats":        stats,
        "findings":     [_f(f) for f in findings_all],
        "assets":       [_a(a) for a in assets],
    }
    try:
        if report_type in ("executive","technical","compliance"):
            path = await PDFReporter().generate(job_id, None, data)
        else:
            path = await ExcelReporter().generate(job_id, None, data)
        return {"message":"리포트 생성 완료","job_id":job_id,
                "download_url":f"/api/report/download/{job_id}"}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"리포트 생성 실패: {str(e)}")


@app.get("/api/report/download/{job_id}")
async def download_report(job_id: str):
    """생성된 리포트 파일 다운로드 (attachment)"""
    base = Path(__file__).parent.parent / "reports"
    for ext in ("pdf", "txt", "xlsx"):
        p = base / f"report_{job_id}.{ext}"
        if p.exists():
            media = "application/pdf" if ext=="pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext=="xlsx" else "text/plain"
            return FileResponse(str(p), filename=f"security_report_{job_id}.{ext}", media_type=media)
    raise HTTPException(404, "리포트 파일을 찾을 수 없습니다. 먼저 생성하세요.")


@app.get("/api/report/view/{job_id}")
async def view_report(job_id: str):
    """생성된 리포트 PDF 브라우저 뷰어로 표시 (inline)"""
    from fastapi.responses import Response
    base = Path(__file__).parent.parent / "reports"
    p = base / f"report_{job_id}.pdf"
    if not p.exists():
        raise HTTPException(404, "리포트 파일을 찾을 수 없습니다.")
    content = p.read_bytes()
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=\"security_report_{job_id}.pdf\""}
    )


@app.post("/api/admin/test-url")
async def test_url(payload: dict):
    """외부 URL 접속 테스트 — 기관 연결 확인용"""
    import requests as _req
    url = payload.get("url","").strip()
    if not url:
        return {"ok":False,"status":None,"latency_ms":None,"error":"URL이 비어 있습니다"}

    def _fetch():
        from datetime import datetime as _dt
        start = _dt.now()
        try:
            r = _req.get(url, timeout=8, verify=False, allow_redirects=True,
                headers={"User-Agent":"SecurityScanKit/1.0"})
            lat = round((_dt.now()-start).total_seconds()*1000)
            return {"ok":r.status_code<400,"status":r.status_code,
                    "latency_ms":lat,"final_url":str(r.url),
                    "error":None if r.status_code<400 else f"HTTP {r.status_code}"}
        except Exception as e:
            lat = round((_dt.now()-start).total_seconds()*1000)
            return {"ok":False,"status":None,"latency_ms":lat,"error":str(e)[:120]}

    # 전역 스레드 풀 사용 (매 요청마다 새 풀 생성 안 함 — Windows 소켓 누수 방지)
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(None, _fetch)
    return result


@app.get("/api/admin/db/fileinfo")
def db_fileinfo():
    """DB 파일 상세 정보 — 경로, 크기, 생성일, 수정일"""
    from db.schema import DB_PATH
    import os
    p = Path(DB_PATH)
    if not p.exists():
        return {"exists": False, "path": str(p)}
    stat = p.stat()
    def fmt_size(b):
        if b < 1024: return f"{b} B"
        if b < 1024**2: return f"{b/1024:.1f} KB"
        return f"{b/1024**2:.2f} MB"
    return {
        "exists":      True,
        "path":        str(p),
        "size_bytes":  stat.st_size,
        "size_human":  fmt_size(stat.st_size),
        "created_at":  datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S"),
        "modified_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "accessed_at": datetime.fromtimestamp(stat.st_atime).strftime("%Y-%m-%d %H:%M:%S"),
    }

@app.get("/api/admin/db/stats")
def db_stats(db: Session = Depends(get_db)):
    """DB 현황 통계"""
    from db.schema import Asset, ScanJob, Finding, Alert, CVERecord, NewsItem, UploadHistory
    return {
        "assets":         db.query(Asset).count(),
        "scan_jobs":      db.query(ScanJob).count(),
        "findings":       db.query(Finding).count(),
        "alerts":         db.query(Alert).count(),
        "cve_records":    db.query(CVERecord).count(),
        "news_items":     db.query(NewsItem).count(),
        "upload_history": db.query(UploadHistory).count(),
    }

class DBResetRequest(BaseModel):
    confirm: str        # "RESET" 입력해야만 실행
    target:  str = "all"  # all / findings / alerts / news / scan_history

@app.post("/api/admin/db/reset")
def db_reset(req: DBResetRequest, db: Session = Depends(get_db)):
    """DB 선택 초기화 — confirm 필드에 RESET 입력 필수"""
    if req.confirm != "RESET":
        raise HTTPException(400, "confirm 필드에 RESET 을 입력하세요")
    from db.schema import (Asset, ScanJob, Finding, Alert,
                           CVERecord, NewsItem, UploadHistory, CVEImpact, Report)
    deleted = {}
    if req.target in ("all","findings"):
        deleted["findings"] = db.query(Finding).delete()
    if req.target in ("all","alerts"):
        deleted["alerts"] = db.query(Alert).delete()
    if req.target in ("all","scan_history"):
        deleted["scan_jobs"] = db.query(ScanJob).delete()
    if req.target in ("all","news"):
        deleted["news_items"] = db.query(NewsItem).delete()
    if req.target == "all":
        db.query(CVEImpact).delete()
        db.query(CVERecord).delete()
        db.query(UploadHistory).delete()
        try: db.query(Report).delete()
        except: pass
        deleted["assets"] = db.query(Asset).delete()
    db.commit()
    if req.target == "all":
        from db.schema import _seed_default_data
        _seed_default_data()
    return {"message": f"DB 초기화 완료 ({req.target})", "deleted": deleted}



# ── WebSocket: 실시간 점검 로그 ────────────────────────────────
@app.websocket("/ws/scan/{job_id}")
async def scan_log_ws(websocket: WebSocket, job_id: str):
    await websocket.accept()
    if job_id not in _ws_clients:
        _ws_clients[job_id] = set()
    _ws_clients[job_id].add(websocket)
    try:
        # 기존 로그 먼저 전송
        if job_id in _scan_logs:
            for entry in list(_scan_logs[job_id]):
                await websocket.send_json(entry)
        # 연결 유지 (ping 수신 대기)
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        pass
    finally:
        if job_id in _ws_clients:
            _ws_clients[job_id].discard(websocket)

@app.get("/api/scan/logs/{job_id}")
def get_scan_logs(job_id: str):
    """WebSocket 불가 환경용 HTTP 폴백"""
    logs = list(_scan_logs.get(job_id, []))
    return {"job_id": job_id, "logs": logs}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=cfg.backend_host, port=cfg.backend_port, reload=True)


# ═══════════════════════════════════════════════════════════════
# 보안 조치 통보 API
# ═══════════════════════════════════════════════════════════════

@app.get("/api/notifications")
def list_notifications(db: Session = Depends(get_db)):
    """통보 이력 전체 조회"""
    rows = db.query(Notification).order_by(Notification.created_at.desc()).limit(200).all()
    return [_notif(n) for n in rows]


@app.get("/api/notifications/summary")
def notification_summary(db: Session = Depends(get_db)):
    """조치 현황 요약 KPI"""
    total     = db.query(Notification).count()
    sent      = db.query(Notification).filter(Notification.status=="sent").count()
    completed = db.query(Notification).filter(Notification.action_status=="completed").count()
    in_prog   = db.query(Notification).filter(Notification.action_status=="in_progress").count()
    overdue   = db.query(Notification).filter(Notification.action_status=="overdue").count()
    notified  = db.query(Notification).filter(Notification.action_status=="notified").count()
    return {"total":total,"sent":sent,"completed":completed,
            "in_progress":in_prog,"overdue":overdue,"notified":notified}


@app.post("/api/notifications/send")
async def send_notification_api(payload: dict, db: Session = Depends(get_db)):
    """담당자 취약점 통보 발송"""
    from notifier import build_manager_pdf, build_html_body, send_notification as _send_mail

    manager    = payload.get("manager","")
    email      = payload.get("email","")
    department = payload.get("department","")
    asset_ids  = payload.get("asset_ids",[])
    due_days   = payload.get("due_days",7)
    smtp       = payload.get("smtp",{})

    if not email:
        raise HTTPException(400, "수신 이메일 주소가 필요합니다")

    # 담당자 자산의 미조치 취약점 조회
    findings_raw = db.query(Finding).join(Asset).filter(
        Asset.id.in_(asset_ids) if asset_ids else Asset.manager==manager,
        Finding.status != "resolved"
    ).all()
    assets_raw = db.query(Asset).filter(
        Asset.id.in_(asset_ids) if asset_ids else Asset.manager==manager
    ).all()

    if not findings_raw:
        raise HTTPException(400, "발송할 취약점이 없습니다 (모두 조치 완료)")

    findings = [_f(f) for f in findings_raw]
    assets   = [_a(a) for a in assets_raw]

    # 담당자 맞춤 PDF 생성
    job_id = f"NOTIF_{str(uuid.uuid4())[:6].upper()}"
    try:
        pdf_path = await build_manager_pdf(manager, findings, assets, job_id)
    except Exception as e:
        pdf_path = None

    # HTML 이메일 본문 생성
    html = build_html_body(manager, department, findings, due_days)
    subject = f"[보안 조치 요청] {manager} 담당 시스템 취약점 {len(findings)}건 — 기한 {due_days}일"

    # 발송
    smtp_cfg = {
        "smtp_host": smtp.get("host",""),
        "smtp_port": int(smtp.get("port",587)),
        "smtp_user": smtp.get("user",""),
        "smtp_pass": smtp.get("password",""),
        "from_addr": smtp.get("from", smtp.get("user","")),
        "use_tls":   smtp.get("use_tls", True),
    }

    result = await _send_mail(
        **smtp_cfg,
        to_email  = email,
        subject   = subject,
        html_body = html,
        pdf_path  = pdf_path,
    )

    # DB 기록
    due_date = datetime.now() + __import__("datetime").timedelta(days=due_days)
    notif = Notification(
        id              = str(uuid.uuid4()),
        manager         = manager,
        manager_email   = email,
        department      = department,
        subject         = subject,
        report_path     = pdf_path or "",
        asset_ids       = json.dumps(asset_ids),
        finding_ids     = json.dumps([f["id"] for f in findings]),
        finding_count   = len(findings),
        critical_count  = sum(1 for f in findings if f.get("severity")=="critical"),
        status          = "sent" if result["ok"] else "failed",
        sent_at         = datetime.now() if result["ok"] else None,
        error_msg       = result.get("error","") if not result["ok"] else "",
        sent_by         = payload.get("sent_by","admin"),
        action_status   = "notified" if result["ok"] else "pending",
        action_due_date = due_date,
    )
    db.add(notif); db.commit(); db.refresh(notif)

    return {
        "ok":      result["ok"],
        "notif_id": notif.id,
        "finding_count": len(findings),
        "pdf_path": pdf_path,
        "error":   result.get("error",""),
    }


@app.patch("/api/notifications/{notif_id}/action")
def update_action_status(notif_id: str, payload: dict, db: Session = Depends(get_db)):
    """조치 상태 업데이트 (담당자 직접 완료 체크)"""
    n = db.query(Notification).filter(Notification.id==notif_id).first()
    if not n: raise HTTPException(404, "통보 이력을 찾을 수 없습니다")

    action = payload.get("action_status")
    if action:
        n.action_status = action
        if action == "completed":
            n.action_completed_at = datetime.now()
            n.action_completed_by = payload.get("completed_by","")
            # 연관 취약점 자동 resolved 처리
            try:
                fids = json.loads(n.finding_ids or "[]")
                if fids:
                    db.query(Finding).filter(Finding.id.in_(fids)).update(
                        {"status":"resolved","resolved_at":datetime.now(),"resolved_by":n.manager},
                        synchronize_session=False
                    )
            except: pass
    if "action_note" in payload:
        n.action_note = payload["action_note"]
    db.commit()
    return _notif(n)


@app.delete("/api/notifications/{notif_id}")
def delete_notification(notif_id: str, db: Session = Depends(get_db)):
    n = db.query(Notification).filter(Notification.id==notif_id).first()
    if not n: raise HTTPException(404)
    db.delete(n); db.commit()
    return {"ok": True}


@app.post("/api/notifications/test-smtp")
async def test_smtp(payload: dict):
    """SMTP 연결 테스트"""
    import smtplib
    host = payload.get("host",""); port = int(payload.get("port",587))
    user = payload.get("user",""); pwd  = payload.get("password","")
    use_tls = payload.get("use_tls", True)
    def _test():
        try:
            if use_tls:
                with smtplib.SMTP(host, port, timeout=8) as s:
                    s.ehlo(); s.starttls()
                    if user: s.login(user, pwd)
            else:
                with smtplib.SMTP_SSL(host, port, timeout=8) as s:
                    if user: s.login(user, pwd)
            return {"ok":True}
        except Exception as e:
            return {"ok":False,"error":str(e)}
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _test)


def _notif(n) -> dict:
    return {
        "id":              n.id,
        "manager":         n.manager,
        "manager_email":   n.manager_email,
        "department":      n.department or "",
        "subject":         n.subject or "",
        "finding_count":   n.finding_count or 0,
        "critical_count":  n.critical_count or 0,
        "status":          n.status,
        "sent_at":         n.sent_at.isoformat() if n.sent_at else None,
        "error_msg":       n.error_msg or "",
        "action_status":   n.action_status,
        "action_due_date": n.action_due_date.isoformat() if n.action_due_date else None,
        "action_note":     n.action_note or "",
        "action_completed_at": n.action_completed_at.isoformat() if n.action_completed_at else None,
        "created_at":      n.created_at.isoformat() if n.created_at else None,
    }


# ─────────────────────────────────────────────────────────────────
# 조치 통보 메일 미리보기 API
# ─────────────────────────────────────────────────────────────────
@app.post("/api/notifications/preview")
async def preview_notification(payload: dict, db: Session = Depends(get_db)):
    """담당자에게 발송될 이메일 HTML 미리보기 생성"""
    from notifier import build_html_body
    manager    = payload.get("manager","")
    department = payload.get("department","")
    asset_ids  = payload.get("asset_ids",[])
    due_days   = payload.get("due_days",7)

    findings_raw = db.query(Finding).join(Asset).filter(
        Asset.id.in_(asset_ids) if asset_ids else Asset.manager==manager,
        Finding.status != "resolved"
    ).all()

    if not findings_raw:
        raise HTTPException(400, "미조치 취약점이 없습니다")

    findings = [_f(f) for f in findings_raw]
    html     = build_html_body(manager, department, findings, due_days)
    return {
        "html": html,
        "finding_count": len(findings),
        "critical_count": sum(1 for f in findings if f.get("severity")=="critical"),
    }


# ═══════════════════════════════════════════════════════════════
# 실시간 로그 API
# ═══════════════════════════════════════════════════════════════
import collections, threading
from fastapi import WebSocket, WebSocketDisconnect

# ════════════════════════════════════════════════════════════════
# 엔터프라이즈 로그 시스템
# ════════════════════════════════════════════════════════════════
_LOG_BUFFER: collections.deque = collections.deque(maxlen=5000)
_LOG_WS_CLIENTS: list = []
_LOG_WS_LOCK = threading.Lock()
_LOG_SEQ = 0          # 단조 증가 시퀀스 번호
_SERVER_START = datetime.now()

# 로그 레벨 필터 (런타임 변경 가능)
_LOG_MIN_LEVEL = "debug"
_LEVEL_ORDER   = {"debug":0, "info":1, "warn":2, "error":3, "critical":4}

# 소스 분류 태그
_SOURCE_TAG = {
    "uvicorn.access":  "ACCESS",
    "uvicorn.error":   "SERVER",
    "uvicorn":         "SERVER",
    "fastapi":         "FASTAPI",
    "sqlalchemy.engine": "SQL",
    "root":            "APP",
}


class _AppLogHandler(logging.Handler):
    """전체 로그를 캡처 → 구조화된 엔트리로 변환 → 버퍼 + WS 브로드캐스트"""
    LEVEL_MAP = {
        "DEBUG":"debug","INFO":"info","WARNING":"warn",
        "ERROR":"error","CRITICAL":"critical"
    }
    def emit(self, record: logging.LogRecord):
        global _LOG_SEQ
        # WS/HTTP 프레임 로그 완전 차단 (무한루프 방지)
        _name = record.name or ""
        if (_name.startswith(("websockets","uvicorn.protocols","h11","httptools","asyncio","concurrent"))
                or "protocols" in _name):
            return
        # > TEXT / < TEXT / > BINARY 등 WS 프레임 패턴 완전 차단
        try:
            _raw = str(record.msg)[:12]
            if (_raw.startswith(("> TEXT", "< TEXT", "> BINA", "< BINA",
                                  "> PING", "< PONG", "> CLOS", "< CLOS",
                                  "= conn", "x half", "connection"))):
                return
            # "bytes]" 패턴도 WS 프레임 로그
            if "bytes]" in str(record.msg) and ("TEXT" in str(record.msg) or "BINARY" in str(record.msg)):
                return
        except Exception:
            return
        level = self.LEVEL_MAP.get(record.levelname, "info")
        if _LEVEL_ORDER.get(level,1) < _LEVEL_ORDER.get(_LOG_MIN_LEVEL,0):
            return

        # 메시지 포맷 (예외 정보 포함)
        try:
            msg = record.getMessage()
        except Exception:
            msg = str(record.msg)

        if record.exc_info:
            import traceback
            msg += "\n" + "".join(traceback.format_exception(*record.exc_info))

        _LOG_SEQ += 1
        entry = {
            "seq":     _LOG_SEQ,
            "ts":      datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "ts_ms":   int(datetime.now().timestamp() * 1000),
            "level":   level,
            "logger":  record.name,
            "tag":     _SOURCE_TAG.get(record.name, record.name.split(".")[-1].upper()[:8]),
            "module":  record.module,
            "func":    record.funcName,
            "line":    record.lineno,
            "thread":  record.thread,
            "msg":     msg,
        }
        _LOG_BUFFER.append(entry)
        # WS 브로드캐스트는 비활성화 (무한루프 방지)
        # 프론트엔드는 /api/logs 폴링 방식으로 조회


# 핸들러 등록
_app_handler = _AppLogHandler()
_app_handler.setFormatter(logging.Formatter("%(message)s"))
_app_handler.setLevel(logging.DEBUG)

_CAPTURE_LOGGERS = [
    "",               # 루트 로거 (print 제외 모든 logging.xxx 캡처)
    "uvicorn",
    "uvicorn.access",
    "uvicorn.error",
    "fastapi",
    "core.intel_collector",
    "reporter",
    "scanner",
    "ssk",
    "__main__",
]
# WS/HTTP 프레임 노이즈 로거 명시적 억제 (CRITICAL로 올려서 emit 자체 차단)
_SUPPRESS_LOGGERS = [
    "websockets", "websockets.server", "websockets.protocol",
    "uvicorn.protocols.websockets", "uvicorn.protocols.websockets.websockets_impl",
    "uvicorn.protocols.websockets.wsproto_impl",
    "h11", "httptools", "uvicorn.protocols.http",
    "uvicorn.protocols.http.httptools_impl",
    "uvicorn.protocols.http.h11_impl",
    "asyncio", "concurrent",
]
for _noisy in _SUPPRESS_LOGGERS:
    _nl = logging.getLogger(_noisy)
    _nl.setLevel(logging.CRITICAL)
    _nl.propagate = False   # ← 루트로 전파 차단

for _ln in _CAPTURE_LOGGERS:
    _l = logging.getLogger(_ln)
    _l.setLevel(logging.DEBUG)
    if not any(isinstance(h, _AppLogHandler) for h in _l.handlers):
        _l.addHandler(_app_handler)

# 앱 자체 로거
app_logger = logging.getLogger("ssk")
app_logger.setLevel(logging.DEBUG)


# ─── REST API ────────────────────────────────────────────────────

@app.get("/api/logs")
def get_logs(
    level:  str = "",
    q:      str = "",
    tag:    str = "",
    logger: str = "",
    after:  int = 0,       # seq 번호 이후 것만 (증분 폴링)
    limit:  int = 1000,
):
    """로그 조회 — 필터 + 증분 폴링 지원"""
    entries = list(_LOG_BUFFER)
    if after:
        entries = [e for e in entries if e["seq"] > after]
    if level and level != "all":
        entries = [e for e in entries if e["level"] == level]
    if q:
        ql = q.lower()
        entries = [e for e in entries if ql in e["msg"].lower()
                   or ql in e["logger"].lower() or ql in e["tag"].lower()]
    if tag:
        entries = [e for e in entries if e["tag"].upper() == tag.upper()]
    if logger:
        entries = [e for e in entries if logger.lower() in e["logger"].lower()]
    return entries[-limit:]


@app.get("/api/logs/stats")
def log_stats():
    """레벨·태그별 통계"""
    from collections import Counter
    entries = list(_LOG_BUFFER)
    lvl_cnt = Counter(e["level"] for e in entries)
    tag_cnt = Counter(e["tag"]   for e in entries)
    # 최근 1분 에러율
    now_ms  = int(datetime.now().timestamp()*1000)
    recent  = [e for e in entries if now_ms - e.get("ts_ms",0) < 60000]
    err_1m  = sum(1 for e in recent if e["level"] in ("error","critical"))
    uptime  = int((datetime.now()-_SERVER_START).total_seconds())
    return {
        "total":     len(entries),
        "seq_max":   _LOG_SEQ,
        "debug":     lvl_cnt.get("debug",0),
        "info":      lvl_cnt.get("info",0),
        "warn":      lvl_cnt.get("warn",0),
        "error":     lvl_cnt.get("error",0),
        "critical":  lvl_cnt.get("critical",0),
        "tags":      dict(tag_cnt.most_common(20)),
        "err_1m":    err_1m,
        "uptime_s":  uptime,
        "min_level": _LOG_MIN_LEVEL,
        "buffer_size": len(_LOG_BUFFER),
    }


@app.patch("/api/logs/level")
def set_log_level(payload: dict):
    """런타임 로그 레벨 변경 — 핸들러 + 캡처 로거 레벨 모두 변경"""
    global _LOG_MIN_LEVEL
    level = payload.get("level","info").lower()
    if level not in _LEVEL_ORDER:
        raise HTTPException(400, f"유효하지 않은 레벨: {level}")
    _LOG_MIN_LEVEL = level
    py_level = getattr(logging, level.upper().replace("WARN","WARNING"))

    # 핸들러 레벨 변경
    _app_handler.setLevel(py_level)

    # 캡처 로거 레벨 변경
    for _ln in _CAPTURE_LOGGERS:
        _l = logging.getLogger(_ln)
        _l.setLevel(py_level)

    app_logger.info(f"[LOG] 로그 레벨 변경: {level.upper()}")
    return {"ok":True,"level":level}


@app.delete("/api/logs")
def clear_logs():
    """로그 버퍼 초기화"""
    global _LOG_SEQ
    _LOG_BUFFER.clear()
    _LOG_SEQ = 0
    return {"ok":True,"cleared":True}


@app.get("/api/logs/export")
def export_logs(level: str = "", fmt: str = "text"):
    """로그 다운로드 (text/json)"""
    from fastapi.responses import PlainTextResponse, Response
    entries = list(_LOG_BUFFER)
    if level and level != "all":
        entries = [e for e in entries if e["level"] == level]
    if fmt == "json":
        import json as _json
        return Response(
            content=_json.dumps(entries, ensure_ascii=False, indent=2),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=ssk_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"}
        )
    lines = [
        f"[{e['ts']}] [{e['level'].upper():8}] [{e['tag']:8}] {e['logger']}:{e['line']} - {e['msg']}"
        for e in entries
    ]
    return PlainTextResponse(
        content="\n".join(lines),
        headers={"Content-Disposition": f"attachment; filename=ssk_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"}
    )


@app.websocket("/ws/logs")
async def ws_logs(websocket: WebSocket):
    """실시간 로그 WebSocket 스트림 — 분할 전송으로 연결 안정화"""
    await websocket.accept()
    with _LOG_WS_LOCK:
        _LOG_WS_CLIENTS.append(websocket)
    try:
        # 최근 200줄을 50개씩 분할 전송 (대용량 bulk 방지)
        recent = list(_LOG_BUFFER)[-200:]
        CHUNK = 50
        for i in range(0, len(recent), CHUNK):
            chunk = recent[i:i+CHUNK]
            await websocket.send_json({"type":"bulk","entries":chunk})
            await asyncio.sleep(0.01)  # 청크 간 소숨

        # keep-alive 루프
        while True:
            try:
                raw = await asyncio.wait_for(websocket.receive_text(), timeout=25)
                try:
                    cmd = __import__("json").loads(raw)
                    if cmd.get("type") == "ping":
                        await websocket.send_json({
                            "type":"pong",
                            "ts":datetime.now().isoformat(),
                            "seq":_LOG_SEQ
                        })
                except Exception:
                    pass
            except asyncio.TimeoutError:
                # heartbeat — 연결 유지
                try:
                    await websocket.send_json({
                        "type":"heartbeat",
                        "ts":datetime.now().isoformat(),
                        "seq":_LOG_SEQ
                    })
                except Exception:
                    break
    except (WebSocketDisconnect, Exception):
        pass
    finally:
        with _LOG_WS_LOCK:
            if websocket in _LOG_WS_CLIENTS:
                _LOG_WS_CLIENTS.remove(websocket)


# ─────────────────────────────────────────────────────────────────
# AI 분석 프록시 API (프론트엔드 CORS 우회)
# ─────────────────────────────────────────────────────────────────
@app.post("/api/ai-analyze")
async def ai_analyze(payload: dict):
    """점검 항목 AI 상세 분석 — Anthropic API 프록시 (CORS 우회)"""
    import requests as _req
    item = payload.get("item", {})
    if not item:
        return {"error": "item 없음"}

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"error": "ANTHROPIC_API_KEY 환경변수가 설정되지 않았습니다. backend/.env 파일에 ANTHROPIC_API_KEY=sk-ant-... 를 추가하세요."}

    prompt = f"""당신은 금융권 보안 전문가입니다. 아래 보안 점검 항목에 대해 JSON 형식으로만 응답하세요. 다른 텍스트나 마크다운 코드블록은 절대 포함하지 마세요.

점검 항목 정보:
- ID: {item.get('id','')}
- 제목: {item.get('title','')}
- 카테고리: {item.get('category','')}
- 심각도: {item.get('severity','')}
- 설명: {item.get('description','')}
- 기준 기관: {item.get('standard','')}{', ' + item.get('standard2','') if item.get('standard2') else ''}
- 기준 근거: {item.get('ref','')}
- 스캐너: {item.get('engine','')}

다음 JSON 구조로만 응답하세요:
{{
  "attack": "공격자가 이 취약점을 실제로 어떻게 악용하는지 구체적인 단계별 시나리오 (3~4문장, 금융권 맥락)",
  "regulation": "금융보안원·금감원·ISMS-P 등 기관에서 이 항목을 어떻게 규정하는지, 위반 시 제재 (3~4문장)",
  "engine": "SecurityScanKit 엔진이 이 취약점을 어떤 기술적 방법으로 탐지하는지 상세 설명 (3~4문장)",
  "cases": "국내외 금융기관에서 실제 발생한 유사 침해 사고 사례 또는 대표적인 공격 사례 (3~4문장)",
  "remediation": "담당자가 즉시 취해야 할 조치 방법을 우선순위 순서로 구체적으로 (번호 목록, 각 1~2문장)",
  "ciso_note": "CISO 관점에서 이 항목의 경영진 보고 시 핵심 메시지 (2~3문장)"
}}"""

    try:
        loop = asyncio.get_event_loop()
        def _call():
            return _req.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "Content-Type": "application/json",
                    "anthropic-version": "2023-06-01",
                    "x-api-key": api_key,
                },
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 1500,
                    "messages": [{"role": "user", "content": prompt}]
                },
                timeout=60, verify=False
            )
        resp = await loop.run_in_executor(None, _call)
        data = resp.json()
        raw = data.get("content", [{}])[0].get("text", "").strip()
        if not raw:
            return {"error": data.get("error", {}).get("message", "빈 응답")}
        # JSON 파싱
        import json as _json
        clean = raw.replace("```json", "").replace("```", "").strip()
        try:
            return {"result": _json.loads(clean)}
        except Exception:
            return {"error": f"JSON 파싱 실패: {clean[:300]}"}
    except Exception as e:
        app_logger.warning(f"[AI-ANALYZE] 오류: {e}")
        return {"error": str(e)}


# 번역 프록시 API (프론트엔드 CORS 우회)
# ─────────────────────────────────────────────────────────────────
@app.post("/api/translate")
async def translate_text(payload: dict):
    """Anthropic API를 프록시해서 영→한 번역 (CORS 우회)"""
    import requests as _req
    text = payload.get("text", "").strip()
    if not text:
        return {"translated": ""}

    # 이미 한글이면 그대로 반환
    kr_cnt = len([ch for ch in text if "가" <= ch <= "힣"])
    if kr_cnt / max(len(text), 1) > 0.25:
        return {"translated": text}

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"translated": text, "error": "ANTHROPIC_API_KEY 환경변수 미설정"}

    try:
        loop = asyncio.get_event_loop()
        def _call():
            return _req.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "Content-Type": "application/json",
                    "anthropic-version": "2023-06-01",
                    "x-api-key": api_key,
                },
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 1000,
                    "messages": [{
                        "role": "user",
                        "content": (
                            "다음 보안 취약점 설명을 자연스러운 한국어로 번역하세요. "
                            "CVE, CVSS, SQL Injection, RCE, XSS 같은 기술 용어는 영어 그대로 유지하세요. "
                            "번역문만 출력하세요.\n\n" + text
                        )
                    }]
                },
                timeout=30, verify=False
            )
        resp = await loop.run_in_executor(None, _call)
        data = resp.json()
        translated = data.get("content", [{}])[0].get("text", "").strip()
        if not translated:
            return {"translated": text, "error": data.get("error", {}).get("message", "빈 응답")}
        return {"translated": translated}
    except Exception as e:
        app_logger.warning(f"[TRANSLATE] 오류: {e}")
        return {"translated": text, "error": str(e)}


# ─────────────────────────────────────────────────────────────────
# 뉴스 관리 API
# ─────────────────────────────────────────────────────────────────
@app.get("/api/news/sources")
def news_sources(db: Session = Depends(get_db)):
    """소스별 뉴스 건수 조회"""
    from db.schema import NewsItem
    from sqlalchemy import func
    rows = db.query(NewsItem.source, func.count(NewsItem.id)).group_by(NewsItem.source).all()
    return [{"source": r[0], "count": r[1]} for r in rows]


@app.delete("/api/news/by-source")
def delete_news_by_source(payload: dict, db: Session = Depends(get_db)):
    """특정 소스의 뉴스 전체 삭제"""
    from db.schema import NewsItem
    sources = payload.get("sources", [])
    if not sources:
        raise HTTPException(400, "삭제할 소스를 지정하세요")
    deleted = 0
    for src in sources:
        cnt = db.query(NewsItem).filter(NewsItem.source == src).delete(synchronize_session=False)
        deleted += cnt
    db.commit()
    app_logger.info(f"[NEWS] 소스 삭제: {sources} — {deleted}건")
    return {"ok": True, "deleted": deleted, "sources": sources}


@app.delete("/api/news/all")
def delete_all_news(db: Session = Depends(get_db)):
    """전체 뉴스 삭제"""
    from db.schema import NewsItem
    cnt = db.query(NewsItem).delete(synchronize_session=False)
    db.commit()
    app_logger.info(f"[NEWS] 전체 삭제: {cnt}건")
    return {"ok": True, "deleted": cnt}


# ─────────────────────────────────────────────────────────────────
# 수집 기관 설정 API
# ─────────────────────────────────────────────────────────────────
# 수집 기관 기본값 — 한국 3개 ON, 해외 OFF
_collect_config: dict = {
    "KrCERT":          True,
    "KISA":            True,
    "금융보안원":       True,
    "CISA":            False,
    "NVD":             True,   # CVE는 NVD만 수집
    "SANS":            False,
    "BleepingComputer":False,
}

@app.get("/api/intel/collect-config")
def get_collect_config():
    """현재 수집 기관 설정 조회"""
    return _collect_config

@app.post("/api/intel/collect-config")
def save_collect_config(payload: dict):
    """수집 기관 ON/OFF 설정 저장"""
    global _collect_config
    _collect_config = payload
    app_logger.info(f"[INTEL] 수집 설정 변경: {payload}")
    return {"ok": True, "config": _collect_config}

# ─────────────────────────────────────────────────────────────────
# 수집 진행 상태 추적
# ─────────────────────────────────────────────────────────────────
_collect_progress = {
    "running":   False,
    "step":      "",        # 현재 단계 메시지
    "steps":     [],        # 전체 단계 로그
    "total":     0,         # 전체 소스 수
    "done":      0,         # 완료된 소스 수
    "saved":     0,         # 저장된 건수
    "started_at": None,
    "finished_at": None,
    "error":     None,
}

@app.get("/api/intel/progress")
def get_collect_progress():
    """수집 진행 상태 조회"""
    return _collect_progress


@app.delete("/api/cve/all")
def delete_all_cve(db: Session = Depends(get_db)):
    """전체 CVE 삭제"""
    from db.schema import CVERecord, CVEImpact
    # 1. 관련 테이블 먼저 삭제
    imp_cnt = db.query(CVEImpact).delete(synchronize_session=False)
    # 2. CVERecord 삭제
    cnt = db.query(CVERecord).delete(synchronize_session=False)
    db.commit()
    # 3. 삭제 후 건수 재확인
    remaining = db.query(CVERecord).count()
    app_logger.info(f"[CVE] 삭제: CVERecord {cnt}건, CVEImpact {imp_cnt}건, 잔여 {remaining}건")
    return {"ok": True, "deleted": cnt, "remaining": remaining}

@app.get("/api/cve/sources")
def cve_sources(db: Session = Depends(get_db)):
    """CVE 소스별 건수"""
    from db.schema import CVERecord
    from sqlalchemy import func, case
    total = db.query(CVERecord).count()
    kev   = db.query(CVERecord).filter(CVERecord.is_kev==True).count()
    nvd   = db.query(CVERecord).filter(CVERecord.is_kev==False).count()
    return [
        {"source": "NVD",      "count": nvd},
        {"source": "CISA KEV", "count": kev},
    ]


# ─────────────────────────────────────────────────────────────────
# 점검 이력 삭제 API
# ─────────────────────────────────────────────────────────────────
@app.delete("/api/scan/history")
def delete_scan_history(payload: dict, db: Session = Depends(get_db)):
    """점검 이력 선택 삭제"""
    from db.schema import ScanJob
    ids = payload.get("ids", [])
    if not ids:
        raise HTTPException(400, "삭제할 이력 ID를 지정하세요")
    cnt = db.query(ScanJob).filter(ScanJob.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    app_logger.info(f"[HISTORY] 점검 이력 삭제: {cnt}건")
    return {"ok": True, "deleted": cnt}

@app.delete("/api/scan/history/all")
def delete_all_scan_history(db: Session = Depends(get_db)):
    """점검 이력 전체 삭제"""
    from db.schema import ScanJob
    cnt = db.query(ScanJob).delete(synchronize_session=False)
    db.commit()
    app_logger.info(f"[HISTORY] 점검 이력 전체 삭제: {cnt}건")
    return {"ok": True, "deleted": cnt}


# ─────────────────────────────────────────────────────────────────
# 클라이언트 IP 반환 API (자산 등록 자동 입력용)
# ─────────────────────────────────────────────────────────────────
@app.get("/api/my-ip")
async def get_my_ip(request: Request):
    """요청자(브라우저) PC의 IP 반환
    - 다른 PC에서 접속하면 그 PC의 IP를 반환
    - 같은 PC면 서버 자신의 IP 반환
    """
    import socket

    def is_private(ip):
        if not ip: return False
        return (ip.startswith("192.168.") or ip.startswith("10.") or
                ip.startswith("172.") or ip.startswith("169.254."))

    private_ips = []

    # 1) 가장 신뢰도 높음 — 요청자 직접 접속 IP
    #    다른 PC에서 접속하면 그 PC의 사설 IP가 여기에 들어옴
    client_ip = request.client.host if request.client else ""
    forwarded = request.headers.get("X-Forwarded-For","").split(",")[0].strip()
    real_ip   = request.headers.get("X-Real-IP","").strip()

    for ip in [forwarded, real_ip, client_ip]:
        if ip and is_private(ip) and ip not in private_ips:
            private_ips.append(ip)

    # 2) 같은 PC(localhost)에서 접속한 경우 — 서버 자신의 IP 수집
    if not private_ips or client_ip in ("127.0.0.1", "::1", ""):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            gw_ip = s.getsockname()[0]
            s.close()
            if is_private(gw_ip) and gw_ip not in private_ips:
                private_ips.insert(0, gw_ip)
        except Exception:
            pass

        try:
            hostname = socket.gethostname()
            addrs = socket.getaddrinfo(hostname, None)
            for addr in addrs:
                ip = addr[4][0]
                if is_private(ip) and ip not in private_ips:
                    private_ips.append(ip)
        except Exception:
            pass

    best_ip = private_ips[0] if private_ips else client_ip
    return {
        "ip": best_ip,
        "candidates": private_ips[:5],
        "client_ip": client_ip,   # 디버그용
        "note": "다른 PC에서 접속 시 해당 PC의 IP가 반환됩니다"
    }


# ── 시스템 사용자 API ─────────────────────────────────────────────

# ── 본부/부서 API ─────────────────────────────────────────────

@app.get("/api/divisions")
def get_divisions(db: Session = Depends(get_db)):
    from db.schema import Division
    rows = db.query(Division).order_by(Division.sort_order, Division.name).all()
    return [{"id": r.id, "name": r.name} for r in rows]

@app.post("/api/divisions")
def add_division(data: dict, db: Session = Depends(get_db)):
    from db.schema import Division
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name required")
    exists = db.query(Division).filter(Division.name == name).first()
    if exists:
        return {"id": exists.id, "name": exists.name}
    row = Division(name=name)
    db.add(row); db.commit(); db.refresh(row)
    return {"id": row.id, "name": row.name}

@app.get("/api/departments")
def get_departments(db: Session = Depends(get_db)):
    from db.schema import Department
    rows = db.query(Department).order_by(Department.sort_order, Department.name).all()
    return [{"id": r.id, "name": r.name, "division_name": r.division_name} for r in rows]

@app.post("/api/departments")
def add_department(data: dict, db: Session = Depends(get_db)):
    from db.schema import Department
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name required")
    exists = db.query(Department).filter(Department.name == name).first()
    if exists:
        return {"id": exists.id, "name": exists.name, "division_name": exists.division_name}
    row = Department(name=name, division_name=(data.get("division_name") or "").strip())
    db.add(row); db.commit(); db.refresh(row)
    return {"id": row.id, "name": row.name, "division_name": row.division_name}


@app.delete("/api/divisions/{div_id}")
def delete_division(div_id: int, db: Session = Depends(get_db)):
    from db.schema import Division
    row = db.query(Division).filter(Division.id == div_id).first()
    if row:
        db.delete(row); db.commit()
    return {"ok": True}

@app.delete("/api/departments/{dept_id}")
def delete_department(dept_id: int, db: Session = Depends(get_db)):
    from db.schema import Department
    row = db.query(Department).filter(Department.id == dept_id).first()
    if row:
        db.delete(row); db.commit()
    return {"ok": True}

@app.get("/api/system-users")
def list_system_users(db: Session = Depends(get_db)):
    from db.schema import SystemUser
    users = db.query(SystemUser).order_by(SystemUser.name).all()
    return [{"id":u.id,"name":u.name,"division":u.division or "","dept":u.dept or "","role":u.role or "user"} for u in users]

@app.post("/api/system-users")
def upsert_system_user(data: dict, db: Session = Depends(get_db)):
    from db.schema import SystemUser
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "이름은 필수입니다")
    u = db.query(SystemUser).filter(SystemUser.name == name).first()
    if u:
        u.division   = data.get("division", u.division)
        u.dept       = data.get("dept", u.dept)
        u.updated_at = datetime.now()
    else:
        u = SystemUser(name=name, division=data.get("division",""), dept=data.get("dept",""), role=data.get("role","user"))
        db.add(u)
    db.commit()
    db.refresh(u)
    return {"id":u.id,"name":u.name,"division":u.division,"dept":u.dept}

@app.post("/api/system-users/bulk")
async def bulk_upsert_system_users(request: Request, db: Session = Depends(get_db)):
    users = await request.json()
    if not isinstance(users, list): users = []
    from db.schema import SystemUser
    saved = 0
    for data in users:
        name = (data.get("name") or "").strip()
        if not name: continue
        u = db.query(SystemUser).filter(SystemUser.name == name).first()
        if u:
            u.division = data.get("division", u.division)
            u.dept     = data.get("dept", u.dept)
            u.role     = data.get("role", u.role)
            u.email    = data.get("email", u.email)
            u.phone    = data.get("phone", u.phone)
        else:
            u = SystemUser(
                name=name,
                division=data.get("division",""),
                dept=data.get("dept",""),
                role=data.get("role","user"),
                email=data.get("email",""),
                phone=data.get("phone",""),
            )
            db.add(u)
        saved += 1
    db.commit()
    return {"saved": saved}

@app.delete("/api/system-users/{user_id}")
def delete_system_user(user_id: int, db: Session = Depends(get_db)):
    from db.schema import SystemUser
    u = db.query(SystemUser).filter(SystemUser.id == user_id).first()
    if not u: raise HTTPException(404, "사용자를 찾을 수 없습니다")
    db.delete(u); db.commit()
    return {"message": "삭제 완료"}

@app.delete("/api/system-users")
def clear_system_users(db: Session = Depends(get_db)):
    from db.schema import SystemUser
    db.query(SystemUser).delete()
    db.commit()
    return {"message": "전체 삭제 완료"}

# ── 프론트엔드 정적 파일 서빙 (Release 모드) ─────────────────────
# npm run build 후 frontend/dist 폴더가 있으면 자동 서빙
from pathlib import Path as _Path
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse as _FileResponse

_DIST = _Path(__file__).parent.parent / "frontend" / "dist"

if _DIST.exists():
    # dist 안의 모든 정적 파일 서빙 (assets, icons 등)
    app.mount("/assets",  StaticFiles(directory=str(_DIST / "assets")),  name="vite-assets")

    # dist 루트에 있는 기타 파일들 (favicon.ico 등)
    @app.get("/favicon.ico", include_in_schema=False)
    async def favicon():
        f = _DIST / "favicon.ico"
        if f.exists(): return _FileResponse(str(f))
        from fastapi import HTTPException; raise HTTPException(404)

    # SPA catch-all: 모든 비-API 경로 → index.html
    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        # API 경로는 건드리지 않음
        if full_path.startswith("api/") or full_path.startswith("docs") or full_path.startswith("openapi"):
            from fastapi import HTTPException; raise HTTPException(status_code=404)
        # 실제 파일이 dist에 있으면 그 파일 서빙
        file_path = _DIST / full_path
        if file_path.exists() and file_path.is_file():
            return _FileResponse(str(file_path))
        # 없으면 SPA index.html 반환 (React Router 처리)
        index = _DIST / "index.html"
        if index.exists():
            return _FileResponse(str(index))
        from fastapi import HTTPException; raise HTTPException(status_code=404)
